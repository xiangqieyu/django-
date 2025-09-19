from django.shortcuts import render,redirect
from app01 import models
from django import forms
from django.core.validators import RegexValidator,ValidationError
from django.utils.safestring import mark_safe
from app01.utils.Pagination import Pagination
from app01.utils.bootstrap import BootStrapModelForm
from app01.utils.encrypt import md5
from openpyxl import load_workbook



def depart_list(request):
    # 部门列表
    # queryset = models.Department.objects.all()
    #
    # return render(request,'depart_list.html', {'queryset': queryset})
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["id__contains"] = search_data

    queryset = models.Department.objects.filter(**data_dict).order_by('-id')

    page_object = Pagination(request, queryset)

    context = {'queryset': page_object.page_queryset,
               'search_data': search_data,
               'page_string': page_object.html()}

    return render(request, "depart_list.html", context)

def depart_add(request):
    # 添加部门
    if request.method == "GET":
        return render(request,"depart_add.html")

    title = request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")

def depart_delete(request):
    # 删除部门
    nid = request.GET.get('id')
    models.Department.objects.filter(id=nid).delete()


    return redirect("/depart/list/")


def depart_edit(request, nid):
    # 修改部门
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id, row_object.title)
        return render(request,"depart_edit.html", {'row_object': row_object})

    title = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")


def depart_multi(request):
    """批量上传（Excel文件）"""
    # 获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 对象传递给openpuxl,由openpuxl读取文件内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect("/depart/list/")


def user_list(request):
    # 用户管理
    # queryset = models.UserInfo.objects.all()
    #
    # return render(request, 'user_list.html', {'queryset': queryset})
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["name__contains"] = search_data

    queryset = models.UserInfo.objects.filter(**data_dict).order_by('-account')

    page_object = Pagination(request, queryset)

    context = {'queryset': page_object.page_queryset,
               'search_data': search_data,
               'page_string': page_object.html()}

    return render(request, "user_list.html", context)

def user_add(request):
    if request.method == "GET":
        context = {
            'gender_choices': models.UserInfo.gender_choices,
            'depart_list': models.Department.objects.all()
        }
        return render(request,'user_add.html',context)
    user = request.POST.get('user')
    pwd = request.POST.get('pwd')
    age = request.POST.get('age')
    ac = request.POST.get('ac')
    ctime = request.POST.get('ctime')
    gd = request.POST.get('gd')
    dp = request.POST.get('dp')

    models.UserInfo.objects.create(name=user,password=pwd,age=age,account=ac,create_time=ctime,gender=gd,depart_id=dp)

    return redirect("/user/list/")

###################################


class UserModelForm(forms.ModelForm):
    class Meta:
        model = models.UserInfo
        widgets = {"create_time": forms.TextInput(attrs={'type': 'date'})}
        fields = ['name', 'password', 'age', 'account', 'create_time', 'gender', 'depart']
        # 定义插件（样式）
        # widgets = {
        #     'name':forms.TextInput(attrs={"class": "form-control"})
        # }
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, field in self.fields.items():
            if field.widget.attrs:
                field.widget.attrs['class'] = "form-control"
                field.widget.attrs['placeholder'] = field.label
            else:
                field.widget.attrs = {"class": "form-control", "placeholder": field.label}
            # field.widget.attrs = {"class": "form-control", "placeholder": field.label}

def user_model_form_add(request):
    if request.method == "GET":
        form = UserModelForm()

        return render(request,'user_model_form_add.html',{'form': form})


    form = UserModelForm(data=request.POST)
    if form.is_valid():

        form.save()
        return redirect("/user/list/")


def user_delete(request):

    nid = request.GET.get('id')
    models.UserInfo.objects.filter(id=nid).delete()

    return redirect('/user/list/')

def user_edit(request, id):
    row_object = models.UserInfo.objects.filter(id=id).first()
    if request.method == "GET":

        form = UserModelForm(instance=row_object)
        return render(request, "user_edit.html", {"form": form})

    form = UserModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/user/list/")
    return render(request, "user_edit.html", {"form": form})


# #########################靓号
def lh_list(request):
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["mobile__contains"] = search_data

    queryset = models.PrettyNum.objects.filter(**data_dict).order_by('-level')

    page_object = Pagination(request, queryset)

    context = {'queryset': page_object.page_queryset,
               'search_data': search_data,
               'page_string': page_object.html()}

    return render(request, "lh_list.html", context)



class LhModelForm(forms.ModelForm):
    #   验证1
    # mobile = forms.CharField(
    #     label='手机号',
    #     validators=[RegexValidator(r'^1\d{10}$', '手机号格式错误')]
    # )
    class Meta:
        model = models.PrettyNum
        fields = ['mobile','price','level','status']
        # fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control", "placeholder": field.label}

    # 验证2
    def clean_mobile(self):
        txt_mobile = self.cleaned_data['mobile']
        exists = models.PrettyNum.objects.filter(mobile=txt_mobile).exists()
        if exists:
            raise ValidationError('手机号已存在')
        # 验证通过
        return txt_mobile





def lh_add(request):
    # 添加靓号
    if request.method == "GET":
        form = LhModelForm()

        return render(request, 'lh_add.html', {'form': form})
    form = LhModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect("/lh/list/")
    return render(request, 'lh_add.html', {'form': form})



class LhEditModelForm(BootStrapModelForm):
    mobile =forms.CharField(disabled=True, label='手机号')
    class Meta:
        model = models.PrettyNum
        fields = ['mobile','price','level','status']
        # fields = "__all__"



    def clean_mobile(self):
        txt_mobile = self.cleaned_data['mobile']
        exists = models.PrettyNum.objects.exclude(id=self.instance.pk).filter(mobile=txt_mobile).exists()
        if exists:
            raise ValidationError('手机号已存在')
        # 验证通过
        return txt_mobile


def lh_edit(request, id):
    row_object = models.PrettyNum.objects.filter(id=id).first()
    if request.method == "GET":
        form = LhEditModelForm(instance=row_object)
        return render(request, "lh_edit.html", {"form": form})

    form = LhEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/lh/list/")
    return render(request, "lh_edit.html", {"form": form})




def lh_delete(request):
    nid = request.GET.get('id')
    models.PrettyNum.objects.filter(id=nid).delete()

    return redirect('/lh/list/')



def admin_list(request):
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["username__contains"] = search_data
    queryset = models.Admin.objects.filter(**data_dict)

    page_object = Pagination(request, queryset)

    context = {'queryset': page_object.page_queryset,
               'search_data': search_data,
               'page_string': page_object.html()}

    return render(request, 'admin_list.html', context)



class AdminModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(label='确认密码', widget=forms.PasswordInput(render_value=True))

    class Meta:
        model = models.Admin
        fields = ['username', 'password', 'confirm_password']
        widgets = {'password': forms.PasswordInput(render_value=True)}


    def clean_password(self):
        pwd = self.cleaned_data.get('password')
        return md5(pwd)

    def clean_confirm_password(self):
        pwd = self.cleaned_data.get('password')
        confirm = md5(self.cleaned_data.get('confirm_password'))
        if confirm != pwd:
            raise ValidationError('密码不一致', code='password_mismatch')
        return confirm


class AdminEditModelForm(BootStrapModelForm):
    class Meta:
        model = models.Admin
        fields = ['username']

class AdminResetModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(label='确认密码', widget=forms.PasswordInput(render_value=True))
    class Meta:
        model = models.Admin
        fields = ['password', 'confirm_password']
        widgets = {'password': forms.PasswordInput(render_value=True)}

    def clean_password(self):
        pwd = self.cleaned_data.get('password')
        md5_pwd = md5(pwd)
        # 去数据库校验当前密码和新输入密码是否一致
        exists = models.Admin.objects.filter(id=self.instance.pk, password=md5_pwd).exists()
        if exists:
            raise ValidationError("密码不能与之前的一致")
        return md5(pwd)

    def clean_confirm_password(self):
        pwd = self.cleaned_data.get('password')
        confirm = md5(self.cleaned_data.get('confirm_password'))
        if confirm != pwd:
            raise ValidationError('密码不一致', code='password_mismatch')
        return confirm



def admin_add(request):
    title = '新建管理员'
    if request.method == 'GET':
        form = AdminModelForm()
        return render(request, 'add.html', {'title': title, 'form': form})

    form = AdminModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')
    print(form.errors)
    return render(request, 'add.html', {'title': title, 'form': form})


def admin_edit(request, id):
    # 编辑管理员
    row_object = models.Admin.objects.filter(id=id).first()
    if not row_object:
        return redirect('/admin/list/')
    title = '编辑管理员'

    if request.method == 'GET':
        form = AdminEditModelForm(instance=row_object)
        return render(request, 'add.html', {'form': form, 'title': title})

    form = AdminEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')

    return render(request, 'add.html', {'form': form, 'title': title})


def admin_delete(request):
    nid = request.GET.get('id')
    models.Admin.objects.filter(id=nid).delete()

    return redirect('/admin/list/')




def admin_reset(request, id):

    row_object = models.Admin.objects.filter(id=id).first()
    if not row_object:
        return redirect('/admin/list/')

    title = "重置密码 - {}".format(row_object.username)

    if request.method == "GET":
        form = AdminResetModelForm(instance=row_object)
        return render(request, 'change.html', {'form': form, 'title': title})

    form = AdminResetModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')

    return render(request, 'change.html', {'form': form, 'title': title})